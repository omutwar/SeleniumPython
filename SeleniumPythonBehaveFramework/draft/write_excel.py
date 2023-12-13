import openpyxl
import os

# Define the file path
# file_path = os.getcwd() # This will get the root path to the script executed
file_path = "..\\..\\resources\\data3.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Sheet2']

sheet.cell(row=1, column=1).value = "Age"
sheet.cell(row=1, column=2).value = "Name"

workbook.save(file_path)
